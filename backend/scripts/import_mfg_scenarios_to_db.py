\
from __future__ import annotations
import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.core import db

MFG_CONFIG_PATH = ROOT / "mock_data" / "mfg" / "config.json"


def load_period_start_dates() -> dict[str, dict[str, date]]:
    if not MFG_CONFIG_PATH.exists():
        return {}
    try:
        cfg = json.loads(MFG_CONFIG_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    result: dict[str, dict[str, date]] = {}
    for region_id, periods in (cfg.get("periods") or {}).items():
        result[region_id] = {}
        for period in periods or []:
            try:
                result[region_id][period["id"]] = date.fromisoformat(period["start_date"])
            except (KeyError, ValueError):
                continue
    return result


def enrich_series_with_dates(series, region_id, period_id, period_starts):
    if not series or any("date" in row for row in series):
        return series
    start = period_starts.get(region_id, {}).get(period_id)
    if not start:
        return series
    return [
        {"date": (start + timedelta(days=index)).isoformat(), **row}
        for index, row in enumerate(series)
    ]

STRATEGY_MAP = {
    "basicscenario": "base",
    "basic": "base",
    "events": "mass_events",
    "maskwearing": "mask_wearing",
    "maskswearing": "mask_wearing",
    "mask": "mask_wearing",
    "totalloc": "full_lockdown",
    "total_loc": "full_lockdown",
    "fulllockdown": "full_lockdown",
    "lockdown": "full_lockdown",
}

def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def slug(value: str) -> str:
    value = value.strip()
    value = re.sub(r"[^0-9A-Za-zА-Яа-я]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value.lower()

def normalize_strategy(source_strategy: str) -> str:
    key = slug(source_strategy).replace("_", "")
    return STRATEGY_MAP.get(key, slug(source_strategy))

def parse_filename(path: Path, default_region_id: str) -> dict[str, str]:
    # Example: SEIR_HCD_TGC_Period1_BasicScenario.xlsx
    name = path.stem
    parts = name.split("_")
    period_part = next((p for p in parts if re.match(r"(?i)^period\d+$", p)), None)
    if period_part:
        period_index = parts.index(period_part)
        source_model = "_".join(parts[: max(period_index - 1, 0)]) or "SEIR_HCD"
        source_region_code = parts[period_index - 1] if period_index >= 1 else default_region_id
        source_period = period_part
        source_strategy = "_".join(parts[period_index + 1:]) or "base"
    else:
        source_model = "unknown"
        source_region_code = default_region_id
        source_period = "period_unknown"
        source_strategy = name
    return {
        "source_model": source_model,
        "source_region_code": source_region_code,
        "source_period": source_period,
        "source_strategy": source_strategy,
        "region_id": default_region_id,
        "period_id": slug(source_period),
        "strategy_id": normalize_strategy(source_strategy),
    }

def to_number_or_string(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == "":
        return None
    normalized = text.replace(",", ".")
    try:
        number = float(normalized)
        return int(number) if number.is_integer() else number
    except ValueError:
        return text

COLUMN_MAP = {
    "time": "t", "day": "t", "t": "t", "date": "date", "дата": "date",
    "s": "S", "susceptible": "S", "e": "E", "exposed": "E", "asymptomatic": "E",
    "i": "I", "infected": "I", "r": "R", "recovered": "R",
    "h": "H", "hospitalized": "H", "c": "C", "critical": "C",
    "d": "D", "dead": "D", "deaths": "D",
}

def normalize_key(key: Any) -> str:
    raw = str(key).strip()
    low = slug(raw).replace("_", "")
    return COLUMN_MAP.get(low, raw)

def normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        item = {}
        for key, value in row.items():
            if key is None:
                continue
            item[normalize_key(key)] = to_number_or_string(value)
        if any(v is not None for v in item.values()):
            normalized.append(item)
    return normalized

def load_json_file(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as f:
        raw = json.load(f)
    if isinstance(raw, list):
        return {"series": normalize_rows(raw), "raw": raw}
    if isinstance(raw, dict):
        if isinstance(raw.get("series"), list):
            raw["series"] = normalize_rows(raw["series"])
            return raw
        for key in ["data", "rows", "results", "values", "Series"]:
            if isinstance(raw.get(key), list):
                return {"series": normalize_rows(raw[key]), "raw": raw}
        for value in raw.values():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                return {"series": normalize_rows(value), "raw": raw}
        return {"series": [], "raw": raw}
    return {"series": [], "raw": raw}

def load_csv_file(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return {"series": normalize_rows(list(reader)), "raw": {"format": "csv"}}

def load_xlsx_file(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return {"series": [], "raw": {"format": "xlsx", "sheet": ws.title}}
    rows = []
    for row in rows_iter:
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return {"series": normalize_rows(rows), "raw": {"format": "xlsx", "sheet": ws.title}}

def load_any_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return load_json_file(path)
    if suffix == ".csv":
        return load_csv_file(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx_file(path)
    raise ValueError(f"Unsupported file type: {path}")

def build_payload(path: Path, meta: dict[str, str], loaded: dict[str, Any], digest: str, period_starts: dict[str, dict[str, date]] | None = None) -> dict[str, Any]:
    scenario = {
        "region_id": meta["region_id"],
        "period_id": meta["period_id"],
        "strategy_id": meta["strategy_id"],
        "source_model": meta["source_model"],
        "source_region_code": meta["source_region_code"],
        "source_period": meta["source_period"],
        "source_strategy": meta["source_strategy"],
    }
    series = loaded.get("series", [])
    if period_starts:
        series = enrich_series_with_dates(
            series, meta["region_id"], meta["period_id"], period_starts
        )
    return {
        "scenario": scenario,
        "model_description": "MFG-SEIR-HCD / статическая модель: данные импортированы из файла.",
        "functional": loaded.get("functional") or loaded.get("cost_functional") or None,
        "parameters": loaded.get("parameters") or [],
        "metrics": [{"id": x, "label": x} for x in ["S", "E", "I", "R", "H", "C", "D"]],
        "series": series,
        "actual_points": loaded.get("actual_points") or [],
        "source": {"type": path.suffix.lower().lstrip("."), "file_name": path.name, "file_path": str(path), "file_hash": digest},
        "raw": loaded.get("raw"),
    }

def main() -> None:
    parser = argparse.ArgumentParser(description="Import only new MFG/static model scenario files into PostgreSQL.")
    parser.add_argument("--input-dir", required=True, help="Folder with JSON/CSV/XLSX scenario files.")
    parser.add_argument("--region-id", default="novosibirsk_oblast", help="Target region_id for imported files.")
    parser.add_argument("--replace", action="store_true", help="Replace existing scenario by scenario_id. Default: skip existing.")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input dir not found: {input_dir}")

    db.init_db(ROOT / "db" / "schema.sql")
    files = []
    for pattern in ("*.json", "*.csv", "*.xlsx", "*.xlsm"):
        files.extend(input_dir.rglob(pattern))

    period_starts = load_period_start_dates()
    imported = 0
    skipped_existing = 0
    failed = 0

    with db.connect() as conn:
        for path in sorted(files):
            stat = path.stat()
            digest = file_sha256(path)
            try:
                meta = parse_filename(path, args.region_id)
                scenario_id = f"{meta['region_id']}:{meta['period_id']}:{meta['strategy_id']}"
                exists = db.mfg_scenario_exists(conn, scenario_id, meta["region_id"], meta["period_id"], meta["strategy_id"], digest)
                if exists and not args.replace:
                    skipped_existing += 1
                    db.record_imported_file(conn, digest, str(path), path.name, stat.st_size, stat.st_mtime, "mfg_scenario", scenario_id, "skipped_existing", "Scenario or file hash already exists in DB")
                    print(f"SKIP existing: {path.name} -> {scenario_id}")
                    continue

                loaded = load_any_file(path)
                payload = build_payload(path, meta, loaded, digest, period_starts)

                if args.replace:
                    db.replace_mfg_scenario(conn, scenario_id, meta["region_id"], meta["period_id"], meta["strategy_id"], payload, meta["source_model"], meta["source_region_code"], meta["source_period"], meta["source_strategy"], str(path), digest)
                    inserted = True
                else:
                    inserted = db.insert_mfg_scenario_if_missing(conn, scenario_id, meta["region_id"], meta["period_id"], meta["strategy_id"], payload, meta["source_model"], meta["source_region_code"], meta["source_period"], meta["source_strategy"], str(path), digest)

                if inserted:
                    imported += 1
                    db.record_imported_file(conn, digest, str(path), path.name, stat.st_size, stat.st_mtime, "mfg_scenario", scenario_id, "imported" if not args.replace else "replaced", f"rows={len(payload.get('series', []))}")
                    print(f"IMPORTED: {path.name} -> {scenario_id} rows={len(payload.get('series', []))}")
                else:
                    skipped_existing += 1
                    print(f"SKIP existing: {path.name} -> {scenario_id}")

            except Exception as exc:
                failed += 1
                db.record_imported_file(conn, digest, str(path), path.name, stat.st_size, stat.st_mtime, "mfg_scenario", None, "failed", str(exc))
                print(f"FAILED: {path.name}: {exc}")

        conn.commit()

    print(f"Done. imported={imported}, skipped_existing={skipped_existing}, failed={failed}")

if __name__ == "__main__":
    main()
